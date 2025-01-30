# OP_Map library: auxiliary functions, model class and its methods
# by Félix Ramón López Martínez
# v1.2
# January-2025
# Release Notes
#   v1.0 first functional code
#   v1.1 added capability to extract forces in crod elements
#        creation of load_cases_list function to optimize the code
#        some functions remaned to distinguish between 2D and 1D forces
#   v1.2 correction of errors associated to read and save excel files
#        and improvement of the code
#

# Import Libraries
import json
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import os
import io

from PIL import Image

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as Img
from openpyxl.styles import Font

from matplotlib.colors import ListedColormap

from pyNastran.op2.op2 import OP2
from pyNastran.bdf.bdf import BDF
from pyNastran.op2.data_in_material_coord import data_in_material_coord


# Definition of OP_Map version
OP_Map_version = '1.2'

########################################################################
# AUXILIARY FUNCTIONS
########################################################################

def read_mapping(mapping_path):
    ''' Reading a json file
        Returning a dictionary with three elements per key:
        + element mapping in numpy format
        + labels for x axis of the mapping
        + labels for the y axis of the mapping
        '''

    # Reading the element mapping dictionary from json file
    with open(mapping_path, 'r') as f:
        mapping = json.load(f)

    # Transforming mapping into numpy darrays
    elm_mapping={}
    for key in mapping.keys():
        elm_mapping[key] = {'mapping': np.array(mapping[key]['mapping']),
                            'x_labels': mapping[key]['x_labels'],
                            'y_labels': mapping[key]['y_labels']}
    
    return elm_mapping


def fig2img(fig):
    """Convert a Matplotlib figure to a PIL Image and return it"""
    buf = io.BytesIO()
    fig.savefig(buf)
    buf.seek(0)
    img = Image.open(buf)
    
    return img


def mapping_extraction(component, mapping_path):
    # Reading mapping from json file
    mapping = read_mapping(mapping_path)

    # Extracting elements mapping for selected component
    elm_mapping = mapping[component]['mapping']
    x_labels = mapping[component]['x_labels']
    y_labels = mapping[component]['y_labels']

    # Extracting fishtail mapping dimensions
    n_dim = elm_mapping.shape[0]
    m_dim = elm_mapping.shape[1]
    # Matrix flattening and turning into a list
    elm_mapping_flt = elm_mapping.reshape(-1,).tolist()
    
    return elm_mapping_flt, x_labels, y_labels, n_dim, m_dim

def formate_axes(plot):
    plot.set_xticklabels(plot.get_xmajorticklabels(), fontsize = 30);
    plot.set_yticklabels(plot.get_ymajorticklabels(), fontsize = 30, rotation = 0);
    ticklbls = plot.get_xticklabels(which='both')    #
    for x in ticklbls:
        x.set_ha('left')

########################################################################
# MODEL CLASSS AND METHODS
########################################################################

class Model:
    def __init__(self, op2_path, mapping_path):
        self.OP_Map_version = OP_Map_version
        self.op2_path = op2_path
        self.mapping_path = mapping_path
        self.op2 = OP2(debug=False)         # instantiate self.op2
        self.elem_to_idx = {}
        self.load_cases = []
        self.lc_list_check = False

        # Creating a new Excel workbook and the Index_of_results sheet
        self.excelfile = os.path.splitext(op2_path)[0] + '.xlsx'   # instantiate excel workbook name
        self.workbook = Workbook()
        self.ws_counter = int(0)                                   # instantiate counter for excel sheets
        ws = self.workbook.active
        ws.sheet_view.showGridLines = False                        # grid lines off
        ws.title = 'Index_of_results'                              # change name of active worksheet
        ws['A1'] = 'Index of results extracted with OP_Map from OP2 file: ' + op2_path
        ws['A1'].font = Font(size = 16, bold = True)
        ws['A2'] = 'OP_Map by Félix R. López M., version: ' + self.OP_Map_version
        ws['A2'].font = Font(size = 8, italic = True)
        self.workbook.save(self.excelfile)                         # Saving the excel workbook

        # Output message
        print('Model initialized with OP_Map library, version', self.OP_Map_version)
        print('Created excel workbook:', self.excelfile)


    def load_cases_list(self, force):
        ''' Method for creating a list with all load cases contained in the OP2
        '''
        if self.lc_list_check == False:
            self.load_cases = [lc for lc in force.keys()]
            print('Number of load cases in the op2 file:', len(self.load_cases)) 
            self.lc_list_check = True
        
        return 


    def r_op2_2D_eforces(self):
        ''' Method for reading the element forces of CTRIA3 y CQUAD4 from the OP2 file
        '''
        # Loading data from OP2
        self.op2.set_results(('force.ctria3_force','force.cquad4_force'))
        self.op2.read_op2(self.op2_path);

        # Getting forces for all subcases  --> diccionary with key the LC and values the force values
        cq4_force = self.op2.cquad4_force
        tr3_force = self.op2.ctria3_force

        # Creating a list with all load cases
        self.load_cases_list(cq4_force)

        # Creating a list with all the elements ID and type
        lc = self.load_cases[0]
        cq4_elements = cq4_force[lc].element
        tr3_elements = tr3_force[lc].element
        elements = np.concatenate((cq4_elements,tr3_elements), axis=0).tolist()

        # Creating a dictionary from element to index starting with 1
        # first index associated to dummy element -666
        self.elem_to_idx[-666]=0
        for idx, elm in enumerate(elements,1):
            self.elem_to_idx[elm] = idx

        # Output message
        print('Loaded element forces of CTRIA3 and CQUAD4 in ELEMENT COORDINATE system from OP2 file:', self.op2_path)
      
        return


    def r_op2_1D_eforces(self):
        ''' Method for reading the element forces of CROD from the OP2 file
        '''
        # Loading data from OP2
        self.op2.set_results(('force.crod_force'))
        self.op2.read_op2(self.op2_path);

        # Getting forces for all subcases  --> diccionary with key the LC and values the force values
        rod_force = self.op2.crod_force

        # Creating a list with all load cases
        self.load_cases_list(rod_force)

        # Creating a list with all the elements ID and type
        lc = self.load_cases[0]
        rod_elements = rod_force[lc].element
        #elements = np.concatenate((cq4_elements,tr3_elements), axis=0).tolist()

        # Creating a dictionary from element to index starting with 1
        # first index associated to dummy element -666
        self.elem_to_idx[-666]=0
        for idx, elm in enumerate(rod_elements,1):
            self.elem_to_idx[elm] = idx

        # Output message
        print('Loaded element forces of CROD elements in ELEMENT COORDINATE system from OP2 file:', self.op2_path)

        return


    def r_op2_2D_eforces_matcoord(self, bdf_path):
        ''' Method for reading the element forces of CTRIA3 y CQUAD4 from the OP2 file
            in material coordinates
            bdf_path is the path to the bdf file that corresponds to the OP2 file
        '''
        # Reading bdf file
        bdf = BDF(debug=False)  # instantiate bdf
        bdf.read_bdf(bdf_path)

        # Creating a new op2 file with 2D results in material coordinates
        self.op2.set_results(('force.ctria3_force','force.cquad4_force'))
        self.op2.read_op2(self.op2_path);
        self.op2 = data_in_material_coord(bdf, self.op2)

        # Getting forces for all subcases  --> diccionary with key the LC and values the force values
        cq4_force = self.op2.cquad4_force
        tr3_force = self.op2.ctria3_force

        # Creating a list with all load cases
        self.load_cases_list(cq4_force)

        # Creating a list with all the elements ID and type
        lc = self.load_cases[0]
        cq4_elements = cq4_force[lc].element
        tr3_elements = tr3_force[lc].element
        elements = np.concatenate((cq4_elements,tr3_elements), axis=0).tolist()

        # Creating a dictionary from element to index starting with 1
        # first index associated to dummy element -666
        self.elem_to_idx[-666]=0
        for idx, elm in enumerate(elements,1):
            self.elem_to_idx[elm] = idx

        # Output message
        print('Loaded element forces of CTRIA3 and CQUAD4 elements in MATERIAL COORDINATE system from OP2 file:', self.op2_path)

        return


    def list_lc(self, excel=False):
        ''' Method for listing the load cases in the OP2 file '''
        print(f'List of load cases contained in the OP2 file: {self.load_cases}')

        if excel:
            self.ws_counter += 1
            tag = 'List of load cases in the OP2 file'
            self.save_in_excel(tag, [self.load_cases])

        return self.load_cases


    def change_mapping(self, new_mapping_path):
        ''' This method change the mapping file for a new one'''
        self.mapping_path = new_mapping_path
        return


    def plot_env_2D_eforces(self, component, env_type, value_field, excel = False):
        ''' This methods plots the MAX, MIN or MAXABS element forces for the
            component and value field passed in.
            Inputs:
                component: name of the component to plot acc. to mapping file
                env_type: MAX, MIN or MAXABS according to the desired envelope
                value_field: force compoment acc. to F06 order
                excel: if True, then results will be stored in the excel workbook
            Output:
                plot of the results (seaborn heatmap image)
                and results saved in the excel file if requested
        '''
        # Mapping extraction for given component
        elm_mapping_flt, x_labels, y_labels, n_dim, m_dim, = mapping_extraction(component, self.mapping_path)

        # Deflattening
        elm_mapping = np.array(elm_mapping_flt).reshape(n_dim, m_dim)

        # Mask creation for filtering results before plotting: 0 for -666 elements and 1 for all the others
        elm_mapping_flt_mask = [1 if elm != -666 else np.nan for elm in elm_mapping_flt]
        elm_mapping_mask = np.array(elm_mapping_flt_mask).reshape(n_dim, m_dim)

        # From element_mapping (flatten) to index_mapping
        # Note that element with id -666 is turn into index 0 (and later in nan when getting the results)
        idx_mapping = [self.elem_to_idx[elm] for elm in elm_mapping_flt]

        # Nunber of load load_cases
        d_dim = len(self.load_cases)

        # Initializing variable output_flt
        output_flt = np.zeros((d_dim, len(idx_mapping)))

        for idx, lc in enumerate(self.load_cases):
            # Accessing individual element forces for a given lC
            cq4_forces_lc = self.op2.cquad4_force[lc].data
            tr3_forces_lc = self.op2.ctria3_force[lc].data

            # Concatenating data from cquads and trias
            forces_lc = np.concatenate((cq4_forces_lc,tr3_forces_lc), axis=1)

            # Removing first dimension
            forces_lc = forces_lc.reshape(forces_lc.shape[1], forces_lc.shape[2])

            # Adding a first line of nan values associated to element index = 0 (dummy element-666)
            nones = np.repeat(np.nan, forces_lc.shape[1], axis=0).reshape(1, forces_lc.shape[1])
            forces_lc = np.concatenate((nones, forces_lc), axis=0)

            # Getting results for the idx_mapping
            output_flt[idx,:] = forces_lc[idx_mapping, (value_field-1)]

        # Deflattening
        output = output_flt.reshape(d_dim, n_dim, m_dim)

        # Envelope options
        if env_type == 'MAX':
            output_env = np.max(output, axis=0)               # Maximum values
            output_env_lc_idx = np.argmax(output, axis=0)     # Index in axis=0 of maximum values (=lc index)
        elif env_type == 'MIN':
            output_env = np.min(output, axis=0)               # Minimum values
            output_env_lc_idx = np.argmin(output, axis=0)     # Index in axis=0 of minimum values (=lc index)
        elif env_type == 'MAXABS':
            output_env = np.max(np.absolute(output), axis=0)  # Maximum absolute values
            output_env_lc_idx = np.argmax(np.absolute(output), axis=0)    # Index in axis=0 of max abs values (=lc index)
        else:
            print('Incorrect envelope type selection. Choose MAX, MIN or MAXABS')

        # From lc_index to lc id
        output_env_lc_flt = [self.load_cases[idx] for idx in output_env_lc_idx.reshape(-1).tolist()]
        output_env_lc = np.array(output_env_lc_flt).reshape(n_dim, m_dim)
        # Removing values from -666 elements
        output_env_lc = np.multiply(output_env_lc, elm_mapping_mask)

        # Plot and Save
        tag = f'Component: {component}. {env_type} element forces in dimension {value_field}'
        self.plot_and_save_env(output_env, output_env_lc, tag, x_labels, y_labels, excel)

        return


    def plot_env_1D_eforces(self, component, env_type, value_field, excel = False):
        ''' This methods plots the MAX, MIN or MAXABS element forces for the
            component and value field passed in.
            Inputs:
                component: name of the component to plot acc. to mapping file
                env_type: MAX, MIN or MAXABS according to the desired envelope
                value_field: force compoment acc. to F06 order
                excel: if True, then results will be stored in the excel workbook
            Output:
                plot of the results (seaborn heatmap image)
                and results saved in the excel file if requested
        '''
        # Mapping extraction for given component
        elm_mapping_flt, x_labels, y_labels, n_dim, m_dim, = mapping_extraction(component, self.mapping_path)

        # Deflattening
        elm_mapping = np.array(elm_mapping_flt).reshape(n_dim, m_dim)

        # Mask creation for filtering results before plotting: 0 for -666 elements and 1 for all the others
        elm_mapping_flt_mask = [1 if elm != -666 else np.nan for elm in elm_mapping_flt]
        elm_mapping_mask = np.array(elm_mapping_flt_mask).reshape(n_dim, m_dim)

        # From element_mapping (flatten) to index_mapping
        # Note that element with id -666 is turn into index 0 (and later in nan when getting the results)
        idx_mapping = [self.elem_to_idx[elm] for elm in elm_mapping_flt]

        # Nunber of load load_cases
        d_dim = len(self.load_cases)

        # Initializing variable output_flt
        output_flt = np.zeros((d_dim, len(idx_mapping)))

        for idx, lc in enumerate(self.load_cases):
            # Accessing individual element forces for a given lC
            rod_forces_lc = self.op2.crod_force[lc].data

            # Concatenating data from cquads and trias
            forces_lc = np.concatenate((rod_forces_lc), axis=1)

            # Adding a first line of nan values associated to element index = 0 (dummy element-666)
            nones = np.repeat(np.nan, forces_lc.shape[1], axis=0).reshape(1, forces_lc.shape[1])
            forces_lc = np.concatenate((nones, forces_lc), axis=0)

            # Getting results for the idx_mapping
            output_flt[idx,:] = forces_lc[idx_mapping, (value_field-1)]

        # Deflattening
        output = output_flt.reshape(d_dim, n_dim, m_dim)
        
        # Envelope options
        if env_type == 'MAX':
            output_env = np.max(output, axis=0)               # Maximum values
            output_env_lc_idx = np.argmax(output, axis=0)     # Index in axis=0 of maximum values (=lc index)
        elif env_type == 'MIN':
            output_env = np.min(output, axis=0)               # Minimum values
            output_env_lc_idx = np.argmin(output, axis=0)     # Index in axis=0 of minimum values (=lc index)
        elif env_type == 'MAXABS':
            output_env = np.max(np.absolute(output), axis=0)  # Maximum absolute values
            output_env_lc_idx = np.argmax(np.absolute(output), axis=0)    # Index in axis=0 of max abs values (=lc index)
        else:
            print('Incorrect envelope type selection. Choose MAX, MIN or MAXABS')
#        # From lc_index to lc id
        output_env_lc_flt = [self.load_cases[idx] for idx in output_env_lc_idx.reshape(-1).tolist()]
        output_env_lc = np.array(output_env_lc_flt).reshape(n_dim, m_dim)
        
        # Removing values from -666 elements
        output_env_lc = np.multiply(output_env_lc, elm_mapping_mask)
        
        # Plot and Save
        tag = f'Component: {component}. {env_type} element forces in dimension {value_field}'
        self.plot_and_save_env(output_env, output_env_lc, tag, x_labels, y_labels, excel)

        return


    def plot_2D_eforces(self, lc, component, value_field, excel = False):
        ''' This methods plots element forces for the load case,
            component and value field passed in.
            Inputs:
                lc: load case to plot
                component: name of the component to plot acc. to mapping file
                value_field: force compoment acc. to F06 order
                excel: if True, then results will be stored in the excel workbook
            Output:
                plot of the results (seaborn heatmap image)
                and results saved in the excel file if requested
        '''
        # Mapping extraction for given component
        elm_mapping_flt, x_labels, y_labels, n_dim, m_dim, = mapping_extraction(component, self.mapping_path)

        # Deflattening
        elm_mapping = np.array(elm_mapping_flt).reshape(n_dim, m_dim)

        # From element_mapping (flatten) to index_mapping
        # Note that element with id -666 is turn into index 0 (and later in nan when getting the results)
        idx_mapping = [self.elem_to_idx[elm] for elm in elm_mapping_flt]

        # Accessing individual element forces for a given lC
        cq4_forces_lc = self.op2.cquad4_force[lc].data
        tr3_forces_lc = self.op2.ctria3_force[lc].data

        # Concatenating data from cquads and trias
        forces_lc = np.concatenate((cq4_forces_lc,tr3_forces_lc), axis=1)

        # Removing first dimension
        forces_lc = forces_lc.reshape(forces_lc.shape[1], forces_lc.shape[2])

        # Adding a first line of nan values associated to element index = 0 (dummy element-666)
        nones = np.repeat(np.nan, forces_lc.shape[1], axis=0).reshape(1, forces_lc.shape[1])
        forces_lc = np.concatenate((nones, forces_lc), axis=0)

        # Getting results for the idx_mapping
        output_flt = forces_lc[idx_mapping, (value_field-1)]

        # Deflattening
        output = output_flt.reshape(n_dim, m_dim)
         
        # Plot and save
        tag = f'Component: {component}. Element forces in dimension {value_field} and for load case {lc}'
        self.plot_and_save(output, tag, x_labels, y_labels, 'coolwarm', excel)

        return


    def plot_1D_eforces(self, lc, component, value_field, excel = False):
        ''' This methods plots element forces for the load case,
            component and value field passed in.
            Inputs:
                lc: load case to plot
                component: name of the component to plot acc. to mapping file
                value_field: force compoment acc. to F06 order
                excel: if True, then results will be stored in the excel workbook
            Output:
                plot of the results (seaborn heatmap image)
                and results saved in the excel file if requested
        '''
        # Mapping extraction for given component
        elm_mapping_flt, x_labels, y_labels, n_dim, m_dim, = mapping_extraction(component, self.mapping_path)

        # Deflattening
        elm_mapping = np.array(elm_mapping_flt).reshape(n_dim, m_dim)

        # From element_mapping (flatten) to index_mapping
        # Note that element with id -666 is turn into index 0 (and later in nan when getting the results)
        idx_mapping = [self.elem_to_idx[elm] for elm in elm_mapping_flt]

        # Accessing individual element forces for a given lC
        rod_forces_lc = self.op2.crod_force[lc].data

        # Concatenating data from cquads and trias
        forces_lc = np.concatenate((rod_forces_lc), axis=1)

        # Adding a first line of nan values associated to element index = 0 (dummy element-666)
        nones = np.repeat(np.nan, forces_lc.shape[1], axis=0).reshape(1, forces_lc.shape[1])
        forces_lc = np.concatenate((nones, forces_lc), axis=0)

        # Getting results for the idx_mapping
        output_flt = forces_lc[idx_mapping, (value_field-1)]

        # Deflattening
        output = output_flt.reshape(n_dim, m_dim)
        
        # Plot and save
        tag = f'Component: {component}. Element forces in dimension {value_field} and for load case {lc}'
        self.plot_and_save(output, tag, x_labels, y_labels, 'coolwarm', excel)

        return


    def plot_component_mapping(self, component, excel = False):
        ''' Method for plotting the mapping of a given component
            Input:
                component
                excel: if True, then results will be stored in the excel workbook
            Output:
                plot of the results (seaborn heatmap image)
                and results saved in the excel file if requested
        '''
        # Mapping extraction for given component
        elm_mapping_flt, x_labels, y_labels, n_dim, m_dim, = mapping_extraction(component, self.mapping_path)

        # From element_mapping (flatten) to index_mapping
        # Note that element with id -666 is turn into index 0 (and later in nan when getting the results)
        elm_mapping_flt = [elm if elm != -666 else np.nan for elm in elm_mapping_flt]

        # Deflattening
        elm_mapping = np.array(elm_mapping_flt).reshape(n_dim, m_dim)

        # Plot and save
        tag = f'Component: {component}. Mapping of elements'
        self.plot_and_save(elm_mapping, tag, x_labels, y_labels, ListedColormap(['whitesmoke']), excel)

        return


    def plot_and_save(self, output, tag, x_labels, y_labels, colormap, excel):
        ''' This methos plot mapings and save them in excel file if requested (excel = True)
        '''
        # Plotting heatmap (fishtail shape)
        plt.figure(figsize=(40,10))
        plot = sns.heatmap(output, annot=True, fmt='.0f',  annot_kws={"size": 20},
                           linewidths=2, cmap=colormap,
                           xticklabels=x_labels, yticklabels=y_labels);
        formate_axes(plot)
        plt.title(tag, fontsize = 30)

        # Turning the plot into an image
        plot_img = fig2img(plot.get_figure())

        # Saving results in the excel workbook if required
        if excel:
            self.ws_counter += 1
            self.save_in_excel(tag, [output], img = plot_img)

        return


    def plot_and_save_env(self, output_env, output_env_lc, tag, x_labels, y_labels, excel):
        ''' This methos plot envelope mapings and save them in excel file if requested (excel = True)
        '''
        # Plotting heatmaps (fishtail shape)
        plt.figure(figsize=(40,20))
        
        # Plot 1
        plt.subplot(2, 1, 1)
        plot_1 = sns.heatmap(output_env, annot=True, fmt='.0f', annot_kws={"size": 20},
                             linewidths=2, cmap='coolwarm',
                             xticklabels=x_labels, yticklabels=y_labels);
        formate_axes(plot_1)
        plt.title(tag, fontsize = 30)

        # Plot 2
        plt.subplot(2, 1, 2)
        plot_2 = sns.heatmap(output_env_lc, annot=True, fmt='.0f', annot_kws={"size": 20},
                             cmap=ListedColormap(['whitesmoke']),
                             linewidths=1, linecolor='White',
                             xticklabels=x_labels, yticklabels=y_labels);
        formate_axes(plot_2)
        plt.title(('Critical load cases. ' + tag), fontsize = 30)

        # Turning the plot into an image
        plot_img = fig2img(plot_1.get_figure())

        # Saving results in the excel workbook if required
        if excel:
            self.ws_counter += 1
            self.save_in_excel(tag, [output_env, output_env_lc], img = plot_img)

        return
    
    
    def save_in_excel(self, tag, data_list, img=None):
        ''' + workbook is the path/name of an already created excel file
            + tag is an identification aboute the data
            + data is a numpy object
            + img is an image (optional)
        '''
        # Load excel workbook
        book = load_workbook(self.excelfile)

        # Name for a new workbook sheet with a sequencial id
        ws_name = f"Sheet_{self.ws_counter}"

        # Write a line in the Index_of_results sheet of the workbook  
        ws_index = book['Index_of_results']
        cell_pos = f'A{4 + self.ws_counter}'
        ws_index[cell_pos] =f"{ws_name} ----> {tag}"
        ws_index[cell_pos].font = Font(size=12, color="000000FF")

        # Hyperlink from the new line in the index to its sheet
        link = '#' + ws_name + '!A1'
        ws_index[cell_pos].hyperlink = link

        # Create new sheet
        ws = book.create_sheet(title=ws_name)

        # Writing the data in the corresponding new sheet
        row_idx = 1
        for data in data_list:
            df = pd.DataFrame(data)    # Create a Pandas dataframe from the data
            for r_idx, row in df.iterrows():
                for c_idx, value in enumerate(row):
                    ws.cell(row=row_idx, column=c_idx+1, value=value)
                row_idx += 1

        # Inserting the image in the new sheet
        if img is not None:
            excel_img = Img(img)
            cell_pos = f"A{row_idx+2}"  # Dejar un espacio antes de la imagen
            ws.add_image(excel_img, cell_pos)

        # Save the workbook and output message
        book.save(self.excelfile)
        print('Results saved in excel workbook:', self.excelfile)

        return